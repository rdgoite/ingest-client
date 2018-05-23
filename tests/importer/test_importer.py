import os
import unittest

from unittest import TestCase

from mock import MagicMock, patch
from openpyxl import Workbook

from ingest.api.ingestapi import IngestApi
from ingest.importer.conversion import template_manager
from ingest.importer.conversion.data_converter import (
    Converter, ListConverter, BooleanConverter, DataType
)
from ingest.importer.data_node import DataNode
from ingest.importer.importer import WorksheetImporter, WorkbookImporter, IngestImporter, SubWorksheetImporter, \
    ProjectWorksheetImporter
from ingest.importer.spreadsheet.ingest_workbook import IngestWorkbook

BASE_PATH = os.path.dirname(__file__)

HEADER_IDX_STR = 4


def _create_single_row_worksheet(worksheet_data: dict):
    workbook = Workbook()
    worksheet = workbook.create_sheet()

    for column, data in worksheet_data.items():
        key, value = data
        worksheet[f'{column}{HEADER_IDX_STR}'] = key
        worksheet[f'{column}6'] = value

    return worksheet


class WorkbookImporterTest(TestCase):

    @patch('ingest.importer.importer.WorksheetImporter')
    def test_do_import(self, worksheet_importer_constructor):
        # given: set up template manager
        key_label_map = {
            'Protocol': 'protocol',
            'Cell Suspension': 'cell_suspension'
        }

        domain_entity_map = {
            'protocol': 'protocol',
            'cell_suspension': 'biomaterial'
        }

        mock_template_manager = MagicMock()
        mock_template_manager.get_concrete_entity_of_tab = lambda key: key_label_map.get(key)
        mock_template_manager.get_domain_entity = lambda key: domain_entity_map.get(key)

        # and: set up worksheet importer
        worksheet_importer = WorksheetImporter()
        expected_json = self._fake_worksheet_import(worksheet_importer, mock_template_manager)

        # and: set up workbook
        workbook = Workbook()
        ingest_workbook = IngestWorkbook(workbook)
        schema_list = self._mock_get_schemas(ingest_workbook)
        self._mock_importable_worksheets(ingest_workbook, workbook)

        # and: mock WorksheetImporter constructor
        worksheet_importer_constructor.return_value = worksheet_importer
        workbook_importer = WorkbookImporter(mock_template_manager)

        # when:
        workbook_output = workbook_importer.do_import(ingest_workbook)

        # then:
        self.assertEqual(2, len(list(workbook_output.keys())))
        self.assertEqual(['protocol', 'biomaterial'], list(workbook_output.keys()))
        self.assertEqual(2, len(list(workbook_output['protocol'].keys())))
        self.assertEqual(expected_json['protocol'], workbook_output['protocol'])
        self.assertEqual(expected_json['biomaterial'], workbook_output['biomaterial'])

    def _mock_get_schemas(self, ingest_workbook):
        schema_base_url = 'https://schema.humancellatlas.org'
        schema_list = [
            f'{schema_base_url}/type/protocol',
            f'{schema_base_url}/type/biomaterial'
        ]
        ingest_workbook.get_schemas = MagicMock(return_value=schema_list)
        return schema_list

    def _mock_importable_worksheets(self, ingest_workbook, workbook):
        project_worksheet = workbook.create_sheet('Protocol')
        cell_suspension_worksheet = workbook.create_sheet('Cell Suspension')
        ingest_workbook.importable_worksheets = MagicMock(return_value=[
            project_worksheet, cell_suspension_worksheet
        ])

    def _fake_worksheet_import(self, worksheet_importer: WorksheetImporter, mock_template_manager):
        projects = {
            'protocol 1': {'short_name': 'protocol 1', 'description': 'first protocol'},
            'protocol 2': {'short_name': 'protocol 2', 'description': 'second protocol'}
        }

        cell_suspensions = {
            'cell_suspension_101': {'biomaterial_id': 'cell_suspension_101', 'biomaterial_name': 'cell suspension'}
        }

        worksheet_iterator = iter([projects, cell_suspensions])
        worksheet_importer.do_import = (
            lambda __, tm: worksheet_iterator.__next__() if tm is mock_template_manager else []
        )

        return {
            'protocol': projects,
            'biomaterial': cell_suspensions
        }


class WorksheetImporterTest(TestCase):

    # TODO refactor this
    def test_do_import(self):
        # given:
        worksheet_importer = WorksheetImporter()

        # and:
        boolean_converter = BooleanConverter()
        converter_mapping = {
            'project.project_core.project_shortname': Converter(),
            'project.miscellaneous': ListConverter(),
            'project.numbers': ListConverter(data_type=DataType.INTEGER),
            'project.is_active': boolean_converter,
            'project.is_submitted': boolean_converter
        }

        concrete_entity_map = {
            'project.project_core.project_shortname': 'project',
            'biomaterial.project_core.project_shortname': 'biomaterial',
        }

        # and:
        mock_template_manager = MagicMock(name='template_manager')
        mock_template_manager.create_template_node = lambda __: DataNode()
        mock_template_manager.get_converter = lambda key: converter_mapping.get(key, Converter())
        mock_template_manager.is_multivalue_object_field = lambda __: False
        mock_template_manager.is_identifier_field = (
            lambda header_name: True if header_name == 'project.project_core.project_shortname' else False
        )
        mock_template_manager.get_concrete_entity_of_column = lambda key: concrete_entity_map.get(key)
        mock_template_manager.get_concrete_entity_of_tab = lambda key: 'project'
        mock_template_manager.get_key_for_label = MagicMock(side_effect=lambda key, tab: key)

        # and:
        worksheet = self._create_test_worksheet()

        # when:
        rows_by_id = worksheet_importer.do_import(worksheet, mock_template_manager)

        # then:
        self.assertEqual(2, len(list(rows_by_id.keys())))
        json = rows_by_id['Tissue stability']['content']

        # and:
        json2 = rows_by_id['Tissue stability 2']['content']
        self.assertEqual('Tissue stability 2', json2['project_core']['project_shortname'])

        project_core = json['project_core']
        self.assertEqual('Tissue stability', project_core['project_shortname'])
        self.assertEqual('Ischaemic sensitivity of human tissue by single cell RNA seq.',
                         project_core['project_title'])

        # and:
        self.assertEqual(2, len(json['miscellaneous']))
        self.assertEqual(['extra', 'details'], json['miscellaneous'])

        # and:
        self.assertEqual(7, json['contributor_count'])

        # and:
        self.assertEqual('Juan Dela Cruz||John Doe', json['contributors'])

        # and:
        self.assertEqual([1, 2, 3], json['numbers'])

        # and:
        self.assertEqual(True, json['is_active'])
        self.assertEqual(False, json['is_submitted'])

    def _create_test_worksheet(self):
        workbook = Workbook()
        worksheet = workbook.create_sheet('Project')
        worksheet[f'A{HEADER_IDX_STR}'] = 'project.project_core.project_shortname'
        worksheet['A6'] = 'Tissue stability'
        worksheet['A7'] = 'Tissue stability 2'
        worksheet[f'B{HEADER_IDX_STR}'] = 'project.project_core.project_title'
        worksheet['B6'] = 'Ischaemic sensitivity of human tissue by single cell RNA seq.'
        worksheet[f'C{HEADER_IDX_STR}'] = 'project.miscellaneous'
        worksheet['C6'] = 'extra||details'
        worksheet[f'D{HEADER_IDX_STR}'] = 'project.contributor_count'
        worksheet['D6'] = 7
        worksheet[f'E{HEADER_IDX_STR}'] = 'project.contributors'
        worksheet['E6'] = 'Juan Dela Cruz||John Doe'
        worksheet[f'F{HEADER_IDX_STR}'] = 'project.numbers'
        worksheet['F6'] = '1||2||3'
        worksheet[f'G{HEADER_IDX_STR}'] = 'project.is_active'
        worksheet['G6'] = 'Yes'
        worksheet[f'H{HEADER_IDX_STR}'] = 'project.is_submitted'
        worksheet['H6'] = 'No'

        return worksheet

    def test_do_import_with_object_list_fields(self):
        # given:
        template_mgr = MagicMock(name='template_manager')
        template_mgr.create_template_node = lambda __: DataNode()
        template_mgr.get_converter = MagicMock(return_value=Converter())
        template_mgr.get_schema_url = MagicMock(return_value='url')
        template_mgr.get_schema_type = MagicMock(return_value='type')
        template_mgr.is_identifier_field = MagicMock(side_effect=(
            lambda header_name: True if header_name == 'project.id_column' else False
        ))
        concrete_entity_map = {
            'project.id_column': 'project'
        }
        template_mgr.get_concrete_entity_of_column = lambda key: concrete_entity_map.get(key)
        template_mgr.get_concrete_entity_of_tab = lambda key: 'project'
        template_mgr.get_key_for_label = MagicMock(side_effect=lambda key, tab: key)

        # and:
        multivalue_fields = {
            'project.genus_species.ontology': True,
            'project.genus_species.text': True,
        }

        template_mgr.is_multivalue_object_field = (
            lambda field_name: multivalue_fields.get(field_name)
        )

        # and:
        worksheet = _create_single_row_worksheet({
            'A': ('project.genus_species.ontology', 'UO:000008'),
            'B': ('project.genus_species.text', 'meter'),
            'C': ('project.id_column', 'id'),
        })

        # and:
        worksheet_importer = WorksheetImporter()

        # when:
        rows_by_id = worksheet_importer.do_import(worksheet, template_mgr)

        # then:
        self.assertEqual(1, len(rows_by_id))
        json = rows_by_id['id']['content']

        # and:
        self.assertTrue(type(json['genus_species']) is list)
        self.assertEqual(1, len(json['genus_species']))
        self.assertEqual({'ontology': 'UO:000008', 'text': 'meter'}, json['genus_species'][0])

    @patch('ingest.importer.importer.ObjectListTracker')
    def test_do_import_builds_from_template(self, object_list_tracker_constructor):
        # given:

        mock_template_manager = MagicMock(name='template_manager')
        mock_template_manager.get_converter = MagicMock(return_value=Converter())
        mock_template_manager.is_multivalue_object_field = MagicMock(return_value=False)
        mock_template_manager.is_identifier_field = (
            lambda header_name: True if header_name == 'project.short_name' else False
        )
        concrete_entity_map = {
            'project.short_name': 'project'
        }
        mock_template_manager.get_concrete_entity_of_column = lambda key: concrete_entity_map.get(key)
        mock_template_manager.get_concrete_entity_of_tab = lambda key: 'project'
        mock_template_manager.get_key_for_label = MagicMock(side_effect=lambda key, tab: key)

        # and:
        node_template = DataNode()
        node_template['describedBy'] = 'https://schemas.sample.com/test'
        node_template['extra_field'] = 'an extra field'
        node_template['version'] = '0.0.1'
        mock_template_manager.create_template_node = lambda __: node_template

        # and:
        object_list_tracker_constructor.return_value = MagicMock(name='object_list_tracker')

        # and:
        importer = WorksheetImporter()

        # and:
        worksheet = _create_single_row_worksheet({
            'A': ('project.short_name', 'Project'),
            'B': ('project.description', 'This is a project')
        })

        # when:
        rows_by_id = importer.do_import(worksheet, mock_template_manager)

        # then:
        self.assertEqual(1, len(list(rows_by_id.keys())))
        json = rows_by_id['Project']['content']

        # and:
        self.assertEqual('https://schemas.sample.com/test', json.get('describedBy'))
        self.assertEqual('an extra field', json.get('extra_field'))
        self.assertEqual('0.0.1', json.get('version'))

    @patch('ingest.importer.importer.ObjectListTracker')
    def test_do_import_with_links(self, object_list_tracker_constructor):
        # given:

        mock_template_manager = MagicMock(name='template_manager')
        mock_template_manager.get_converter = MagicMock(return_value=Converter())
        mock_template_manager.is_multivalue_object_field = MagicMock(return_value=False)

        identifier_field_map = {
            'project.short_name': True,
            'cell_suspension.cell_suspension_id': True
        }

        mock_template_manager.is_identifier_field = (
            lambda header_name: identifier_field_map.get(header_name)
        )
        concrete_entity_map = {
            'project.short_name': 'project',
            'cell_suspension.cell_suspension_id': 'cell_suspension'
        }
        mock_template_manager.get_concrete_entity_of_column = lambda key: concrete_entity_map.get(key)
        mock_template_manager.get_concrete_entity_of_tab = lambda key: 'project'

        domain_entity_map = {
            'project': 'project',
            'cell_suspension': 'biomaterial'
        }

        mock_template_manager.get_domain_entity = lambda concrete_entity: domain_entity_map.get(concrete_entity)

        mock_template_manager.get_key_for_label = MagicMock(side_effect=lambda key, tab: key)

        # and:
        node_template = DataNode()
        node_template['describedBy'] = 'https://schemas.sample.com/test'
        node_template['extra_field'] = 'an extra field'
        node_template['version'] = '0.0.1'
        mock_template_manager.create_template_node = lambda __: node_template

        # and:
        object_list_tracker_constructor.return_value = MagicMock(name='object_list_tracker')

        # and:
        importer = WorksheetImporter()

        # and:
        worksheet = _create_single_row_worksheet({
            'A': ('project.short_name', 'Project id'),
            'B': ('project.description', 'This is a project'),
            'C': ('cell_suspension.cell_suspension_id', 'cell_suspension_id')
        })

        # when:
        rows_by_id = importer.do_import(worksheet, mock_template_manager)

        # then:
        self.assertEqual(1, len(list(rows_by_id.keys())))
        links = rows_by_id['Project id']['links_by_entity']
        content = rows_by_id['Project id']['content']

        # and:
        self.assertTrue(links.get('biomaterial'), 'Must have a link to a biomaterial')
        self.assertEqual(1, len(links.get('biomaterial')))
        self.assertEqual('cell_suspension_id', links.get('biomaterial')[0])
        self.assertFalse(content.get('cell_suspension_id'), 'Must not contain link field value in the content object.')


class SubWorksheetImporterTest(TestCase):

    def test_do_import(self):
        worksheet = _create_single_row_worksheet({
            'A': ('project.contributors.contact_name', 'Contributor Name'),
            'B': ('project.contributors.email', 'contributor@project.ac.uk'),
            'C': ('project.contributors.phone', '1234567')
        })
        mock_template_manager = MagicMock(name='template_manager')
        mock_template_manager.get_converter = MagicMock(return_value=Converter())
        mock_template_manager.is_multivalue_object_field = MagicMock(return_value=True)
        node_template = DataNode()
        node_template['describedBy'] = 'https://schemas.sample.com/test'
        node_template['extra_field'] = 'an extra field'
        node_template['version'] = '0.0.1'
        mock_template_manager.create_template_node = lambda __: node_template

        importer = SubWorksheetImporter()
        rows = importer.do_import(worksheet, mock_template_manager)

        self.assertEqual(1, len(rows))
        self.assertEqual('Contributor Name', rows[0]['contact_name'])
        self.assertEqual('contributor@project.ac.uk', rows[0]['email'])
        self.assertEqual('1234567', rows[0]['phone'])


class ProjectWorksheetImporterTest(TestCase):

    def test_do_import(self):
        worksheet = _create_single_row_worksheet({
            'A': ('project.path.title', 'Project Title'),
            'B': ('project.path.description', 'Project Description'),
            'C': ('project.path.name', 'Project Name')
        })
        mock_template_manager = MagicMock(name='template_manager')
        mock_template_manager.get_converter = MagicMock(return_value=Converter())
        mock_template_manager.is_multivalue_object_field = MagicMock(return_value=False)
        node_template = DataNode()
        node_template['describedBy'] = 'https://schemas.sample.com/test'
        node_template['extra_field'] = 'an extra field'
        node_template['version'] = '0.0.1'
        mock_template_manager.create_template_node = lambda __: node_template

        importer = ProjectWorksheetImporter()
        project = importer.do_import(worksheet, mock_template_manager)

        self.assertEqual('Project Title', project['path']['title'])
        self.assertEqual('Project Description', project['path']['description'])
        self.assertEqual('Project Name', project['path']['name'])


class IngestImporterTest(TestCase):

    @unittest.skip
    def test_import_spreadsheet(self):
        ingest_api = IngestApi(url='http://localhost:8080')
        submission_url = ingest_api.createSubmission(None)
        spreadsheet_file = BASE_PATH + '/metadata_spleen_new_protocols.xlsx'
        submission = IngestImporter(ingest_api).import_spreadsheet(file_path=spreadsheet_file, submission_url=submission_url, dry_run=False)

        self.assertTrue(submission)
